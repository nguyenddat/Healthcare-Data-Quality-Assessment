from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..completeness.schemas.overview import TableOverviewInfo
from .utils import write_sheet

OVERVIEW_HEADER_LABELS = {
    "ten_bang": "Tên bảng",
    "so_dong": "Số dòng thực tế",
    "ngay_tao": "Ngày tạo",
    "du_lieu": "Dữ liệu",
    "tier": "Tier",
}

DU_LIEU_FILL = {
    "no_data": "FFC7CE",
    "sparse_data": "FFEB9C",
    "full_data": "C6EFCE",
}

def build_overview_workbook(rows: list[TableOverviewInfo]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    write_sheet(
        ws,
        OVERVIEW_HEADER_LABELS,
        [row.model_dump() for row in rows],
        fill_map=DU_LIEU_FILL,
        fill_field="du_lieu",
    )
    return wb


def extract_tiered_table_names(ws: Worksheet) -> list[str]:
    header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    ten_bang_col = header[OVERVIEW_HEADER_LABELS["ten_bang"]]
    tier_col = header[OVERVIEW_HEADER_LABELS["tier"]]
    names = []
    for row in ws.iter_rows(min_row=2):
        tier_value = row[tier_col - 1].value
        if tier_value is not None and str(tier_value).strip() != "":
            names.append(row[ten_bang_col - 1].value)
    return names
