from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL_COLOR = "305496"
HEADER_FONT_COLOR = "FFFFFF"
BORDER_COLOR = "B7B7B7"


def write_sheet(
    ws: Worksheet,
    header_labels: dict[str, str],
    rows: list[dict],
    fill_map: dict[str, str] | None = None,
    fill_field: str | None = None,
) -> None:
    fields = list(header_labels.keys())

    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, field in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_labels[field])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, data in enumerate(rows, start=2):
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=data[field])
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        if fill_map and fill_field:
            fill_col = fields.index(fill_field) + 1
            fill = PatternFill("solid", fgColor=fill_map.get(data[fill_field], "FFFFFF"))
            ws.cell(row=row_idx, column=fill_col).fill = fill

    for col_idx, field in enumerate(fields, start=1):
        max_len = max(
            [len(header_labels[field])]
            + [len(str(data[field])) for data in rows]
            + [1]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
